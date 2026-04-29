import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date
from fpdf import FPDF
import os

st.set_page_config(
    page_title="MR Fakturagenerator (Ajour + Dansk Omsorgspleje + Dit Vikarbureau)",
    layout="centered",
)

# ---------- Styling ----------
st.markdown(
    """
<style>
body { background-color:#aa1e1e; color:white; }
[data-testid="stAppViewContainer"] > .main {
    background:white;
    color:black;
    border-radius:10px;
    padding:2rem;
    max-width:980px;
    margin:auto;
}
</style>
""",
    unsafe_allow_html=True,
)

if os.path.exists("logo.png"):
    st.image("logo.png", width=80)

# --------------------------------------------------
# CONSTANTS (FROM / TO)
# --------------------------------------------------
FROM_INFO = {
    "name": "MR Rekruttering",
    "addr": "Valbygårdsvej 1, 4. th, 2500 Valby",
    "cvr": "45090965",
    "phone": "71747290",
    "web": "www.akutvikar.com",
}

BANK_INFO_LINE1 = "Bank: Finseta | IBAN: GB79TCCL04140404627601 | BIC: TCCLGB3LXXX"
BANK_INFO_LINE2 = "Betalingsbetingelser: Bankoverførsel. Fakturanr. bedes angivet ved betaling."

TO_AJOUR = {
    "title": "Ajour Care ApS",
    "lines": [
        "CVR: 34478953",
        "Kontakt: Charlotte Bigum Christensen",
        "Email: cbc@ajourcare.dk",
    ],
}

TO_DANSK = {
    "title": "DANSK OMSORGSPLEJE APS",
    "lines": [
        "CVR: 42092630",
        "Frederiksborgvej 14, st, 3200 Helsinge",
    ],
}

TO_DIT = {
    "title": "Dit Vikarbureau",
    "lines": [
        "CVR: (indsæt CVR)",
        "Adresse: (indsæt adresse)",
        "Email: (indsæt email)",
    ],
}

# --------------------------------------------------
# HELPERS
# --------------------------------------------------
def normalize_personale(val: str) -> str:
    if val is None:
        return ""
    s = str(val).replace("\u00A0", " ").strip().lower()
    s = " ".join(s.split())
    if s == "assistent 2":
        s = "assistent"
    if "ufagl" in s:
        return "ufaglært"
    if "hjælp" in s:
        return "hjælper"
    if "assist" in s:
        return "assistent"
    if "sygepl" in s:
        return "sygeplejerske"
    if "ergoter" in s:
        return "ergoterapeut"
    return s


def ensure_datetime(series) -> pd.Series:
    return pd.to_datetime(series, dayfirst=True, errors="coerce")


def safe_time_str(val) -> str:
    if pd.isna(val):
        return ""
    s = str(val)
    if len(s) >= 5 and s[2] == ":":
        return s[:5]
    return s[:5]


def build_tidsperiode(start, end) -> str:
    return f"{safe_time_str(start)}-{safe_time_str(end)}"


def parse_start_time_to_minutes(tidsperiode: str) -> int:
    try:
        s = str(tidsperiode).split("-")[0].strip()
        hh, mm = s.split(":")
        return int(hh) * 60 + int(mm)
    except Exception:
        return 0


def time_to_hour(t: str) -> int:
    try:
        return int(str(t)[:2])
    except Exception:
        return 0


# --------------------------------------------------
# BASE CLEANING (used for all)
# --------------------------------------------------
def rens_data_base(df: pd.DataFrame) -> pd.DataFrame:
    needed = [
        "Dato",
        "Medarbejder",
        "Starttid",
        "Sluttid",
        "Timer",
        "Personalegruppe",
        "Jobfunktion",
        "Shift status",
        "Afdeling",
    ]
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise ValueError(f"Mangler kolonner i filen: {', '.join(missing)}")

    df = df[needed].copy()

    df["Timer"] = pd.to_numeric(df["Timer"], errors="coerce")
    df = df[df["Timer"].notna() & (df["Timer"] > 0)].copy()

    df["Dato"] = ensure_datetime(df["Dato"])
    df = df[df["Dato"].notna()].copy()

    df["Tidsperiode"] = df.apply(lambda r: build_tidsperiode(r["Starttid"], r["Sluttid"]), axis=1)
    df["Jobfunktion_raw"] = df["Jobfunktion"]

    df["Personale"] = df["Personalegruppe"].apply(normalize_personale)

    df["StartMin"] = df["Tidsperiode"].apply(parse_start_time_to_minutes)

    return df


# --------------------------------------------------
# AJOURCARE: jobfunktion mapping
# --------------------------------------------------
def map_jobfunktion_ajour(df: pd.DataFrame) -> pd.DataFrame:
    byer = ["allerød", "egedal", "frederiksund", "frederikssund", "solrød", "herlev", "ringsted", "køge"]

    def find_by(txt):
        t = str(txt).lower()
        if "ergoter" in t:
            return "ergoterapeut"
        for b in byer:
            if b in t:
                if b in ("frederiksund", "frederikssund"):
                    return "frederikssund"
                return b
        if "kirsten" in t:
            return "køge"
        return "andet"

    out = df.copy()
    out["Jobfunktion"] = out["Jobfunktion"].apply(find_by)
    return out


# --------------------------------------------------
# DANSK OMSORGSPLEJE: jobfunktion display
# --------------------------------------------------
def extract_location_dansk(jobfunction):
    if not jobfunction:
        return ""
    parts = str(jobfunction).split("-")
    return parts[1].strip() if len(parts) > 1 else str(jobfunction).strip()


def extract_location_dit(jobfunction):
    return extract_location_dansk(jobfunction)


# --------------------------------------------------
# RATE LOGIC
# --------------------------------------------------
def beregn_takst_ajour(row) -> float:
    helligdag = row["Helligdag"] == "Ja"
    personale = row["Personale"]

    start_hour = time_to_hour(row["Tidsperiode"].split("-")[0])
    dag = start_hour < 15
    weekend = row["Dato"].weekday() >= 5

    if personale == "ufaglært":
        if helligdag:
            return 215 if dag else 220
        return 215 if weekend and dag else 220 if weekend else 175 if dag else 210

    if personale == "hjælper":
        if helligdag:
            return 215 if dag else 220
        return 215 if weekend and dag else 220 if weekend else 200 if dag else 210

    if personale == "assistent":
        if helligdag:
            return 230 if dag else 240
        return 230 if weekend and dag else 240 if weekend else 220 if dag else 225

    if personale == "sygeplejerske":
        if helligdag:
            return 695 if dag else 790
        return 520 if weekend and dag else 615 if weekend else 370 if dag else 465

    if personale == "ergoterapeut":
        dag = start_hour < 15
        if helligdag:
            return 700 if dag else 790
        return 400 if dag else 480

    return 0


def beregn_takst_dansk(row) -> float:
    if row["Helligdag"] == "Ja":
        return 350

    weekend = row["Dato"].weekday() >= 5
    if weekend:
        return 300

    start_hour = time_to_hour(row["Tidsperiode"].split("-")[0])
    return 280 if start_hour >= 15 else 255


DITVIKAR_RATES = {
    "hjælper": {
        "weekday_day": 333.00,
        "weekday_night": 406.26,
        "weekend_day": 499.50,
        "weekend_night": 572.76,
        "holiday_day": 666.00,
        "holiday_night": 739.26,
    },
    "assistent": {
        "weekday_day": 353.00,
        "weekday_night": 430.66,
        "weekend_day": 529.50,
        "weekend_night": 607.16,
        "holiday_day": 706.00,
        "holiday_night": 783.66,
    },
    "sygeplejerske": {
        "weekday_day": 386.00,
        "weekday_night": 482.50,
        "weekend_day": 579.00,
        "weekend_night": 675.50,
        "holiday_day": 772.00,
        "holiday_night": 868.50,
    },
}


def is_day_window(start_min: int) -> bool:
    return 6 * 60 <= start_min < 15 * 60


def beregn_takst_dit(row) -> float:
    personale = row["Personale"]
    rates = DITVIKAR_RATES.get(personale)
    if not rates:
        return 0.0

    helligdag = row["Helligdag"] == "Ja"
    weekend = row["Dato"].weekday() >= 5
    start_min = parse_start_time_to_minutes(row["Tidsperiode"])
    day = is_day_window(start_min)

    if helligdag:
        return rates["holiday_day"] if day else rates["holiday_night"]
    if weekend:
        return rates["weekend_day"] if day else rates["weekend_night"]
    return rates["weekday_day"] if day else rates["weekday_night"]


# --------------------------------------------------
# EXCEL GENERATION
# --------------------------------------------------
def generer_excel(invoices: dict) -> BytesIO:
    """invoices = {sheet_name: (inv_df, fakturanr, to_info)}"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    RED = "AA1E1E"
    LIGHT_RED = "F5CCCC"
    GREY = "F2F2F2"

    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name, (inv, fakturanr, to_info) in invoices.items():
        ws = wb.create_sheet(title=sheet_name[:31])

        # Header block
        ws.merge_cells("A1:I1")
        ws["A1"] = f"FAKTURA {fakturanr}  –  {FROM_INFO['name']}"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=RED)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:I2")
        ws["A2"] = (
            f"Fra: {FROM_INFO['addr']}  |  CVR: {FROM_INFO['cvr']}  |  "
            f"Tlf: {FROM_INFO['phone']}  |  Web: {FROM_INFO['web']}"
        )
        ws["A2"].font = Font(size=9, color="555555")
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:I3")
        ws["A3"] = f"Til: {to_info['title']}  |  " + "  |  ".join(to_info["lines"])
        ws["A3"].font = Font(size=9)
        ws["A3"].alignment = Alignment(horizontal="center")
        ws["A3"].fill = PatternFill("solid", fgColor=LIGHT_RED)

        ws.merge_cells("A4:I4")
        ws["A4"] = f"Fakturadato: {date.today().strftime('%d.%m.%Y')}"
        ws["A4"].font = Font(size=9, italic=True)
        ws["A4"].alignment = Alignment(horizontal="right")

        ws.append([])  # blank row 5

        # Column headers row 6
        headers = ["Dato", "Medarbejder", "Tidsperiode", "Timer", "Personale", "Jobfunktion", "Helligdag", "Takst (kr)", "Samlet (kr)"]
        col_widths = [14, 28, 16, 8, 16, 28, 10, 12, 14]
        ws.append(headers)
        for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=6, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=RED)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.row_dimensions[6].height = 20

        # Data rows
        data_start = 7
        for i, (_, r) in enumerate(inv.iterrows()):
            row_num = data_start + i
            fill = PatternFill("solid", fgColor=GREY) if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            row_data = [
                r["Dato"].strftime("%d.%m.%Y"),
                str(r["Medarbejder"]),
                r["Tidsperiode"],
                float(r["Timer"]),
                str(r["Personale"]),
                str(r["Jobfunktion"]),
                r["Helligdag"],
                float(r["Takst"]),
                float(r["Samlet"]),
            ]
            ws.append(row_data)
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = fill
                cell.border = border
                cell.font = Font(size=9)
                if col_idx in [1, 3, 5, 6, 7]:
                    cell.alignment = Alignment(horizontal="center")
                if col_idx == 4:
                    cell.number_format = "0.0"
                    cell.alignment = Alignment(horizontal="center")
                if col_idx in [8, 9]:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

        data_end = data_start + len(inv) - 1

        # Totals
        total_row = data_end + 2
        ws.cell(row=total_row, column=7).value = "Subtotal:"
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=9).value = f"=SUM(I{data_start}:I{data_end})"
        ws.cell(row=total_row, column=9).number_format = "#,##0.00"
        ws.cell(row=total_row, column=9).font = Font(bold=True)

        moms_row = total_row + 1
        ws.cell(row=moms_row, column=7).value = "Moms (25%):"
        ws.cell(row=moms_row, column=7).font = Font(bold=True)
        ws.cell(row=moms_row, column=7).alignment = Alignment(horizontal="right")
        ws.cell(row=moms_row, column=9).value = f"=I{total_row}*0.25"
        ws.cell(row=moms_row, column=9).number_format = "#,##0.00"
        ws.cell(row=moms_row, column=9).font = Font(bold=True)

        grand_row = moms_row + 1
        ws.cell(row=grand_row, column=7).value = "Total inkl. moms:"
        ws.cell(row=grand_row, column=7).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=grand_row, column=7).fill = PatternFill("solid", fgColor=RED)
        ws.cell(row=grand_row, column=7).alignment = Alignment(horizontal="right")
        ws.cell(row=grand_row, column=9).value = f"=I{total_row}+I{moms_row}"
        ws.cell(row=grand_row, column=9).number_format = "#,##0.00"
        ws.cell(row=grand_row, column=9).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=grand_row, column=9).fill = PatternFill("solid", fgColor=RED)

        # Footer
        footer_row = grand_row + 2
        ws.merge_cells(f"A{footer_row}:I{footer_row}")
        ws.cell(row=footer_row, column=1).value = BANK_INFO_LINE1
        ws.cell(row=footer_row, column=1).font = Font(size=8, italic=True)
        ws.merge_cells(f"A{footer_row+1}:I{footer_row+1}")
        ws.cell(row=footer_row + 1, column=1).value = BANK_INFO_LINE2
        ws.cell(row=footer_row + 1, column=1).font = Font(size=8, italic=True)

        ws.freeze_panes = "A7"
        ws.sheet_view.showGridLines = False

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# --------------------------------------------------
# PDF GENERATION (generic)
# --------------------------------------------------
def generer_pdf(inv: pd.DataFrame, fakturanr: int, to_info: dict, filename_prefix: str) -> tuple[BytesIO, str]:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=18)

    if os.path.exists("logo.png"):
        pdf.image("logo.png", 10, 5, 30)

    pdf.set_font("Arial", "B", 20)
    pdf.set_xy(140, 10)
    pdf.cell(60, 10, f"FAKTURA {fakturanr}", align="R")

    pdf.set_font("Arial", "B", 12)
    pdf.set_xy(10, 40)
    pdf.cell(95, 6, f"Fra: {FROM_INFO['name']}", ln=1)
    pdf.set_font("Arial", "", 10)
    pdf.set_x(10)
    pdf.cell(95, 6, FROM_INFO["addr"], ln=1)
    pdf.set_x(10)
    pdf.cell(95, 6, f"CVR.nr. {FROM_INFO['cvr']}", ln=1)
    pdf.set_x(10)
    pdf.cell(95, 6, f"Tlf: {FROM_INFO['phone']}", ln=1)
    pdf.set_x(10)
    pdf.cell(95, 6, f"Web: {FROM_INFO['web']}", ln=1)
    y_left_end = pdf.get_y()

    pdf.set_font("Arial", "B", 12)
    pdf.set_xy(105, 40)
    pdf.cell(95, 6, f"Til: {to_info['title']}", ln=1)
    pdf.set_font("Arial", "", 10)
    for line in to_info["lines"]:
        pdf.set_x(105)
        pdf.cell(95, 6, line, ln=1)
    y_right_end = pdf.get_y()

    header_end_y = max(y_left_end, y_right_end)
    pdf.set_y(header_end_y + 6)

    pdf.set_x(10)
    pdf.cell(0, 6, f"Fakturadato: {date.today().strftime('%d.%m.%Y')}", ln=1)
    pdf.ln(4)

    cols = ["Dato", "Medarbejder", "Tidsperiode", "Timer", "Personale", "Jobfunktion", "Helligdag", "Takst", "Samlet"]
    widths = [18, 28, 20, 10, 22, 32, 12, 14, 14]

    pdf.set_font("Arial", "B", 8)
    pdf.set_x(10)
    for h, w in zip(cols, widths):
        pdf.cell(w, 8, h, 1, align="C")
    pdf.ln()

    pdf.set_font("Arial", "", 8)
    total = 0.0
    for _, r in inv.iterrows():
        pdf.set_x(10)

        row = [
            r["Dato"].strftime("%d.%m.%Y"),
            str(r["Medarbejder"])[:26],
            r["Tidsperiode"],
            f"{float(r['Timer']):.1f}",
            str(r["Personale"]),
            str(r["Jobfunktion"])[:40],
            r["Helligdag"],
            f"{float(r['Takst']):.2f}",
            f"{float(r['Samlet']):.2f}",
        ]

        for i, (v, w) in enumerate(zip(row, widths)):
            align = "C"
            if i in [1, 4, 5]:
                align = "L"
            pdf.cell(w, 8, str(v), 1, align=align)

        pdf.ln()
        total += float(r["Samlet"])

    moms = total * 0.25
    pdf.ln(4)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, f"Subtotal: {total:.2f} kr", ln=1)
    pdf.cell(0, 6, f"Moms (25%): {moms:.2f} kr", ln=1)
    pdf.cell(0, 6, f"Total inkl. moms: {total + moms:.2f} kr", ln=1)

    pdf.ln(5)
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 6, BANK_INFO_LINE1, ln=1)
    pdf.cell(0, 6, BANK_INFO_LINE2, ln=1)

    pdf_bytes = pdf.output(dest="S").encode("latin-1")
    out = BytesIO(pdf_bytes)
    out.seek(0)
    return out, f"{filename_prefix}_{fakturanr}.pdf"


# --------------------------------------------------
# BUILD INVOICE DF
# --------------------------------------------------
def build_invoice_df(df_customer: pd.DataFrame, helligdage: list, rate_func) -> pd.DataFrame:
    inv = df_customer.copy()

    hellig_set = set(pd.to_datetime(helligdage))
    inv["Helligdag"] = inv["Dato"].dt.normalize().isin([h.normalize() for h in hellig_set]).map({True: "Ja", False: "Nej"})

    inv["Takst"] = [rate_func(r) for _, r in inv.iterrows()]
    inv["Samlet"] = inv["Timer"] * inv["Takst"]

    inv = inv[
        ["Dato", "Medarbejder", "Tidsperiode", "StartMin", "Timer", "Personale", "Jobfunktion", "Helligdag", "Takst", "Samlet"]
    ].copy()

    inv = inv.sort_values(["Jobfunktion", "Dato", "StartMin", "Medarbejder"], ascending=[True, True, True, True])
    inv = inv.drop(columns=["StartMin"])
    return inv


# --------------------------------------------------
# UI
# --------------------------------------------------
st.title("MR Rekruttering – Fakturagenerator (Ajour Care + Dansk Omsorgspleje + Dit Vikarbureau)")

file = st.file_uploader("Upload vagtplan-fil (Excel)", type=["xlsx", "xls"])

c1, c2, c3 = st.columns(3)
with c1:
    faktura_ajour = st.number_input("Fakturanummer (Ajour Care / AkutVikar)", min_value=0, step=1, value=0)
with c2:
    faktura_dansk = st.number_input("Fakturanummer (Dansk Omsorgsplejle)", min_value=0, step=1, value=0)
with c3:
    faktura_dit = st.number_input("Fakturanummer (Dit Vikarbureau)", min_value=0, step=1, value=0)

# Export format toggle
st.markdown("### Eksportformat")
export_format = st.radio(
    "Vælg filformat",
    options=["PDF", "Excel (.xlsx)", "Begge"],
    horizontal=True,
)

if file:
    try:
        raw = pd.read_excel(file)
        clean = rens_data_base(raw)

        afdeling_lower = clean["Afdeling"].astype(str).str.lower()

        df_ajour = clean[
            afdeling_lower.str.contains(r"ajour|akut\s*-?\s*vikar|akutvikar", regex=True, na=False)
        ].copy()

        df_dansk = clean[
            afdeling_lower.str.contains(r"dansk\s*omsorgspleje", regex=True, na=False)
        ].copy()

        df_dit = clean[
            afdeling_lower.str.contains(r"dit\s*vikar|ditvikar|dit\s*vikarbureau|dit\s*vikarbuerou", regex=True, na=False)
        ].copy()

        if len(df_ajour) > 0:
            df_ajour = map_jobfunktion_ajour(df_ajour)

        if len(df_dansk) > 0:
            df_dansk["Jobfunktion"] = df_dansk["Jobfunktion"].apply(extract_location_dansk)

        if len(df_dit) > 0:
            df_dit["Jobfunktion"] = df_dit["Jobfunktion"].apply(extract_location_dit)

        st.markdown("### Fil-status")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Rækker (total)", len(clean))
        m2.metric("Ajour/Akut Vikar rækker", len(df_ajour))
        m3.metric("Dansk Omsorgspleje rækker", len(df_dansk))
        m4.metric("Dit Vikarbureau rækker", len(df_dit))

        all_dates = sorted(clean["Dato"].dt.date.unique())
        helligdage = [pd.Timestamp(d) for d in st.multiselect("Vælg helligdage (gælder for alle)", all_dates)]

        can_generate = True
        if len(df_ajour) > 0 and faktura_ajour <= 0:
            can_generate = False
            st.warning("Du har Ajour/Akut Vikar-rækker, men mangler fakturanummer til Ajour/AkutVikar.")
        if len(df_dansk) > 0 and faktura_dansk <= 0:
            can_generate = False
            st.warning("Du har Dansk Omsorgspleje-rækker, men mangler fakturanummer til Dansk Omsorgspleje.")
        if len(df_dit) > 0 and faktura_dit <= 0:
            can_generate = False
            st.warning('Du har Dit Vikarbureau-rækker, men mangler fakturanummer til "Dit Vikarbureau".')

        if st.button("Generer fakturaer", disabled=not can_generate):
            st.success("Fakturaer genereres…")

            inv_ajour = inv_dansk = inv_dit = None

            # Build invoice DataFrames
            if len(df_ajour) > 0:
                inv_ajour = build_invoice_df(df_ajour, helligdage, beregn_takst_ajour)

                # Kirsten +10
                kirsten_flag = df_ajour[["Dato", "Medarbejder", "Tidsperiode", "Jobfunktion_raw"]].copy()
                kirsten_flag["Kirsten"] = kirsten_flag["Jobfunktion_raw"].astype(str).str.contains("kirsten", case=False, na=False)
                kirsten_flag = kirsten_flag.drop(columns=["Jobfunktion_raw"])

                inv_ajour = inv_ajour.merge(kirsten_flag, on=["Dato", "Medarbejder", "Tidsperiode"], how="left")
                inv_ajour["Kirsten"] = inv_ajour["Kirsten"].fillna(False)
                inv_ajour.loc[inv_ajour["Kirsten"] == True, "Takst"] = inv_ajour["Takst"] + 10
                inv_ajour["Samlet"] = inv_ajour["Timer"] * inv_ajour["Takst"]
                inv_ajour = inv_ajour.drop(columns=["Kirsten"])

            if len(df_dansk) > 0:
                inv_dansk = build_invoice_df(df_dansk, helligdage, beregn_takst_dansk)

            if len(df_dit) > 0:
                inv_dit = build_invoice_df(df_dit, helligdage, beregn_takst_dit)

            if inv_ajour is None and inv_dansk is None and inv_dit is None:
                st.error("Ingen rækker fundet for Ajour/Akut Vikar, Dansk Omsorgspleje eller Dit Vikarbureau i Afdeling-kolonnen.")
            else:
                # --- PDF downloads ---
                if export_format in ("PDF", "Begge"):
                    if inv_ajour is not None:
                        pdf_ajour, pdf_ajour_name = generer_pdf(inv_ajour, int(faktura_ajour), TO_AJOUR, "Faktura_AjourCare")
                        st.download_button("📄 Download PDF (Ajour/AkutVikar)", pdf_ajour, file_name=pdf_ajour_name)

                    if inv_dansk is not None:
                        pdf_dansk, pdf_dansk_name = generer_pdf(inv_dansk, int(faktura_dansk), TO_DANSK, "Faktura_DanskOmsorgspleje")
                        st.download_button("📄 Download PDF (Dansk Omsorgspleje)", pdf_dansk, file_name=pdf_dansk_name)

                    if inv_dit is not None:
                        pdf_dit, pdf_dit_name = generer_pdf(inv_dit, int(faktura_dit), TO_DIT, "Faktura_DitVikarbureau")
                        st.download_button("📄 Download PDF (Dit Vikarbureau)", pdf_dit, file_name=pdf_dit_name)

                # --- Excel downloads (separate per customer) ---
                if export_format in ("Excel (.xlsx)", "Begge"):
                    if inv_ajour is not None:
                        excel_ajour = generer_excel({"Ajour Care": (inv_ajour, int(faktura_ajour), TO_AJOUR)})
                        st.download_button(
                            "📊 Download Excel (Ajour/AkutVikar)",
                            excel_ajour,
                            file_name=f"Faktura_AjourCare_{int(faktura_ajour)}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

                    if inv_dansk is not None:
                        excel_dansk = generer_excel({"Dansk Omsorgspleje": (inv_dansk, int(faktura_dansk), TO_DANSK)})
                        st.download_button(
                            "📊 Download Excel (Dansk Omsorgspleje)",
                            excel_dansk,
                            file_name=f"Faktura_DanskOmsorgspleje_{int(faktura_dansk)}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

                    if inv_dit is not None:
                        excel_dit = generer_excel({"Dit Vikarbureau": (inv_dit, int(faktura_dit), TO_DIT)})
                        st.download_button(
                            "📊 Download Excel (Dit Vikarbureau)",
                            excel_dit,
                            file_name=f"Faktura_DitVikarbureau_{int(faktura_dit)}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

    except Exception as e:
        st.error(f"Fejl: {e}")
